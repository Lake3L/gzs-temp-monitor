"""
Подсистема формирования отчётов и уведомлений.
Генерирует отчёты в различных форматах и отправляет уведомления.
"""

import os
from datetime import datetime, timedelta
from typing import Optional, List
from dataclasses import dataclass
from enum import Enum

from database.db_manager import DatabaseManager
from database.models import Event, EventType, EventStatus


class ReportPeriod(Enum):
    """Периоды отчётов"""
    DAILY = "За день"
    WEEKLY = "За неделю"
    MONTHLY = "За месяц"
    CUSTOM = "Произвольный"


class ExportFormat(Enum):
    """Форматы экспорта"""
    XLSX = "Excel (.xlsx)"
    PDF = "PDF (.pdf)"


@dataclass
class ReportData:
    """Данные отчёта"""
    title: str
    period_start: datetime
    period_end: datetime
    generated_at: datetime
    events: List[Event]
    statistics: dict
    summary: str


class Reporter:
    """
    Генератор отчётов и уведомлений.
    """
    
    def __init__(self, db_manager: DatabaseManager, 
                 reports_dir: str = "reports"):
        """
        Инициализация генератора отчётов.
        
        Args:
            db_manager: Менеджер базы данных
            reports_dir: Директория для сохранения отчётов
        """
        self.db = db_manager
        self.reports_dir = reports_dir
        
        # Создаём директорию для отчётов
        if not os.path.exists(reports_dir):
            os.makedirs(reports_dir)
    
    def generate_report(self, period: ReportPeriod,
                        start_date: datetime = None,
                        end_date: datetime = None,
                        room_id: int = None,
                        event_type: EventType = None) -> ReportData:
        """
        Сгенерировать данные отчёта.
        
        Args:
            period: Период отчёта
            start_date: Начало периода (для CUSTOM)
            end_date: Конец периода (для CUSTOM)
            room_id: Фильтр по помещению
            event_type: Фильтр по типу события
            
        Returns:
            ReportData с данными отчёта
        """
        # Определяем период
        end_date = end_date or datetime.now()
        
        if period == ReportPeriod.DAILY:
            start_date = end_date - timedelta(days=1)
            title = f"Ежедневный отчёт за {end_date.strftime('%d.%m.%Y')}"
        elif period == ReportPeriod.WEEKLY:
            start_date = end_date - timedelta(weeks=1)
            title = f"Еженедельный отчёт ({start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')})"
        elif period == ReportPeriod.MONTHLY:
            start_date = end_date - timedelta(days=30)
            title = f"Ежемесячный отчёт ({start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')})"
        else:
            if not start_date:
                start_date = end_date - timedelta(days=7)
            title = f"Отчёт за период ({start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')})"
        
        # Получаем события за период
        events = self.db.get_events(
            start_time=start_date,
            end_time=end_date,
            room_id=room_id,
            event_type=event_type
        )
        
        # Получаем статистику
        statistics = self.db.get_statistics(start_date, end_date)
        
        # Формируем сводку
        summary = self._generate_summary(events, statistics)
        
        return ReportData(
            title=title,
            period_start=start_date,
            period_end=end_date,
            generated_at=datetime.now(),
            events=events,
            statistics=statistics,
            summary=summary
        )
    
    def _generate_summary(self, events: List[Event], statistics: dict) -> str:
        """Генерация текстовой сводки"""
        total_events = len(events)
        warnings = sum(1 for e in events if e.event_type == EventType.WARNING)
        critical = sum(1 for e in events if e.event_type == EventType.CRITICAL)
        failures = sum(1 for e in events if e.event_type == EventType.SENSOR_FAILURE)
        
        resolved = sum(1 for e in events if e.status == EventStatus.RESOLVED)
        active = sum(1 for e in events if e.status == EventStatus.ACTIVE)
        
        lines = [
            f"Всего событий: {total_events}",
            f"  - Предупреждений: {warnings}",
            f"  - Критических: {critical}",
            f"  - Сбоев датчиков: {failures}",
            "",
            f"Разрешено: {resolved}",
            f"Активных: {active}",
        ]
        
        return "\n".join(lines)
    
    def export_xlsx(self, report_data: ReportData, 
                    filename: str = None) -> str:
        """
        Экспортировать отчёт в Excel.
        
        Args:
            report_data: Данные отчёта
            filename: Имя файла (если не указано, генерируется автоматически)
            
        Returns:
            Путь к созданному файлу
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise ImportError("Для экспорта в Excel необходим пакет openpyxl")
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"report_{timestamp}.xlsx"
        
        filepath = os.path.join(self.reports_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчёт по событиям"
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок отчёта
        ws.merge_cells('A1:G1')
        ws['A1'] = report_data.title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_align
        
        ws.merge_cells('A2:G2')
        ws['A2'] = f"Период: {report_data.period_start.strftime('%d.%m.%Y %H:%M')} - {report_data.period_end.strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].alignment = center_align
        
        ws.merge_cells('A3:G3')
        ws['A3'] = f"Сформирован: {report_data.generated_at.strftime('%d.%m.%Y %H:%M:%S')}"
        ws['A3'].alignment = center_align
        
        # Заголовки таблицы
        headers = ["№", "Дата/Время", "Датчик", "Тип события", "Температура", "Статус", "Примечания"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Данные
        for row_idx, event in enumerate(report_data.events, 6):
            sensor = self.db.get_sensor(event.sensor_id)
            sensor_name = sensor.name if sensor else f"Датчик #{event.sensor_id}"
            
            row_data = [
                row_idx - 5,
                event.start_time.strftime('%d.%m.%Y %H:%M'),
                sensor_name,
                event.event_type.value,
                f"{event.temperature:.1f}°C",
                event.status.value,
                event.notes or ""
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                if col in [1, 4, 5, 6]:
                    cell.alignment = center_align
        
        # Ширина колонок
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 30
        
        # Лист статистики
        ws_stats = wb.create_sheet("Статистика")
        
        ws_stats['A1'] = "Статистика по событиям"
        ws_stats['A1'].font = Font(bold=True, size=12)
        
        stats_row = 3
        for line in report_data.summary.split('\n'):
            ws_stats.cell(row=stats_row, column=1, value=line)
            stats_row += 1
        
        wb.save(filepath)
        return filepath
    
    def export_pdf(self, report_data: ReportData,
                   filename: str = None) -> str:
        """
        Экспортировать отчёт в PDF.
        
        Args:
            report_data: Данные отчёта
            filename: Имя файла
            
        Returns:
            Путь к созданному файлу
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            raise ImportError("Для экспорта в PDF необходим пакет reportlab")
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"report_{timestamp}.pdf"
        
        filepath = os.path.join(self.reports_dir, filename)
        
        # Регистрация шрифта с поддержкой кириллицы
        try:
            # Пробуем использовать Arial
            pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
            font_name = 'Arial'
        except:
            try:
                # Альтернатива - DejaVuSans
                pdfmetrics.registerFont(TTFont('DejaVuSans', 'DejaVuSans.ttf'))
                font_name = 'DejaVuSans'
            except:
                font_name = 'Helvetica'
        
        doc = SimpleDocTemplate(
            filepath,
            pagesize=landscape(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1*cm,
            bottomMargin=1*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Заголовок
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontName=font_name,
            fontSize=16,
            alignment=1,
            spaceAfter=12
        )
        elements.append(Paragraph(report_data.title, title_style))
        
        # Период
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=10,
            alignment=1,
            spaceAfter=20
        )
        period_text = f"Период: {report_data.period_start.strftime('%d.%m.%Y %H:%M')} - {report_data.period_end.strftime('%d.%m.%Y %H:%M')}"
        elements.append(Paragraph(period_text, subtitle_style))
        
        # Таблица событий
        table_data = [["№", "Дата/Время", "Датчик", "Тип", "Темп.", "Статус"]]
        
        for idx, event in enumerate(report_data.events[:50], 1):  # Ограничиваем 50 записями
            sensor = self.db.get_sensor(event.sensor_id)
            sensor_name = sensor.name if sensor else f"#{event.sensor_id}"
            
            table_data.append([
                str(idx),
                event.start_time.strftime('%d.%m.%Y %H:%M'),
                sensor_name[:20],
                event.event_type.value[:15],
                f"{event.temperature:.1f}°C",
                event.status.value[:12]
            ])
        
        table = Table(table_data, colWidths=[1*cm, 4*cm, 5*cm, 4*cm, 2.5*cm, 3*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # Сводка
        summary_style = ParagraphStyle(
            'Summary',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=10,
            leading=14
        )
        summary_text = report_data.summary.replace('\n', '<br/>')
        elements.append(Paragraph(f"<b>Сводка:</b><br/>{summary_text}", summary_style))
        
        doc.build(elements)
        return filepath
    
    def send_notification(self, event: Event, recipients: List[str] = None):
        """
        Отправить уведомление о событии.
        
        Args:
            event: Событие
            recipients: Список получателей email
        """
        # Получаем настройки уведомлений
        if not recipients:
            settings = self.db.get_notification_settings(event_type=event.event_type)
            recipients = [s.email for s in settings]
        
        if not recipients:
            return
        
        sensor = self.db.get_sensor(event.sensor_id)
        sensor_name = sensor.name if sensor else f"Датчик #{event.sensor_id}"
        
        room = self.db.get_room(sensor.room_id) if sensor else None
        room_name = room.name if room else "Неизвестно"
        
        subject = f"[{event.event_type.value}] {sensor_name}"
        message = (
            f"Тип события: {event.event_type.value}\n"
            f"Датчик: {sensor_name}\n"
            f"Помещение: {room_name}\n"
            f"Температура: {event.temperature:.1f}°C\n"
            f"Порог: {event.threshold_exceeded:.1f}°C\n"
            f"Время: {event.start_time.strftime('%d.%m.%Y %H:%M:%S')}\n"
        )
        
        for recipient in recipients:
            # Симуляция отправки (в реальном приложении здесь был бы SMTP)
            self._send_email(recipient, subject, message)
            
            # Логируем уведомление
            self.db.add_notification(
                event_id=event.id,
                channel="email",
                recipient=recipient,
                message=message
            )
    
    def _send_email(self, recipient: str, subject: str, message: str):
        """
        Отправка email (симуляция).
        В реальном приложении здесь был бы SMTP клиент.
        """
        print(f"[EMAIL] To: {recipient}")
        print(f"[EMAIL] Subject: {subject}")
        print(f"[EMAIL] Message: {message[:100]}...")
